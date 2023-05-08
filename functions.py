import openpyxl

'''
Project specific functions 

Convert column letter to number
Compiling comments into one space
Outputting into specified format (xlsx, sql, etc)
'''

def getCorrectPath():
    '''
    Assist the user in getting the right file path
    '''

    better_path = input("Enter a new path to xlsx workbook: ")
    return better_path

def getWorkbook(path: str):
    '''
    Verifies that the given path works and creates a workbook object
    Then creates a sheet object which is used throughout the whole process

    :param path: Path to xlsx workbook
    '''

    # Catches any issues related to the path and attempts to handle them before moving on
    # Instead of looping endlessly, it gives the user a chance to adjust the path of the workbook
    try:
        wb_obj = openpyxl.load_workbook(path)

    except FileNotFoundError:
        print("\nThe path is incorrect or the file does not exist")
        new_path = getCorrectPath()
        wb_obj = openpyxl.load_workbook(new_path)

    except:
        print("\nThere was an issue with the given path")
        new_path = getCorrectPath()
        wb_obj = openpyxl.load_workbook(new_path)

    # Returns a working sheet for the program to use
    sheet = wb_obj.active
    return sheet

def workbookGrid(sheet: openpyxl.workbook.workbook.Workbook):
    '''
    Creates a grid of the current sheet for use
    Using a nested list

    :param sheet: The active sheet obj to create a grid for
    '''


    # Each list corresponds to a ROW
    # Each value on the row corresponds to a COLUMN
 
    tot_row = sheet.max_row
    tot_column = sheet.max_column 
    sheetGrid = []


    # Nested list comprehension
    # The first row/list corresponds to the questions
    # The following rows/list corresponds to the responses

    for list in range(tot_row + 1):
        sheetGrid.append([])
        for item in range(tot_column + 1):
            print(item)
            sheetGrid[list].append(sheet.cell(row = list + 1, column = item + 1).value)

    return sheetGrid

def getColumnInfo(sheet: openpyxl.workbook.workbook.Workbook, col: int, ):
    '''
    Get all the values from a specific column 
    And returns a list with those values 

    :param sheet: The active sheet obj to get info from 
    :param col: Column to get values from 
    ''' 

    values = []

    # Looping through the column and appending each value after the column name to a list
    for i in range(sheet.max_row):

        # Skips the first iteration and prints the column name
        if i == 0:
            print(f"Column: {sheet.cell(row = 1, column = col).value}")
            continue
        
        # Prints the current value and appends to list
        # Comment out unless needed
        #print(sheet.cell(row = i + 1, column = col).value)

        value = sheet.cell(row = i + 1, column = col).value
        
        # If you want to make values lowercase
        # if type(value) == str:
        #    value = value.lower()

        values.append(value)

    return values

def getCorrespondingRows(sheet: openpyxl.workbook.workbook.Workbook, col: int, search_term):
    '''
    Iterates through the workbook with the specified search term
    And returns a list of the row numbers that match that information based on the column given

    Note to self...
    Merging with corresponding rows using a dictionary to compress the code might not be a bad idea,

    :param sheet: The active sheet obj to get info from 
    :param col: Column number iterate though
    :param search: Value used when iterating through column to find the match
    '''

    values = getColumnInfo(sheet, col)
    matched_rows = []

    # If the search term is a word, this makes everything lowercase for comparion purposes
    if type(search_term) == str:
        
        count = 0
        search_term = search_term.lower()
        
        for value in values:
            values[count] = value.lower()
            count += 1

    # Starts at the second row sinces first row is reserved for column names
    row_num = 2 

    # Iterates through the column list and appends the row number of the matched value
    for value in values: 
        if value == search_term:
            print("Match found")
            matched_rows.append(row_num)
        row_num += 1

    return matched_rows

def getAverageForColumn(sheet: openpyxl.workbook.workbook.Workbook, col: int):
    '''
    Returns the average of the values in the given column

    :param sheet: The active sheet obj to get info from 
    :param col: Column to calculate the average values from
    '''

    values = getColumnInfo(sheet, col)
    total = 0
    count = 0

    num_or_letter = int(input("Enter 1 for number based averaging or 2 for letter based averaging (not working now...): "))
    
    # For number based averaging
    if num_or_letter == 1:
        num = 0       

        # Confirms that values are ints
        for value in values:
            values[num] = int(value)
            num += 1

        for item in values:
            total += item
            count += 1

        average = total/count

        return average
    
    # Comment out string based for production testing
    # For string based averaging
    elif num_or_letter == 2:
        criteria = [input("Enter the grading criteria from the highest rating to the lowest, seperated by space\n(I.E. : Excellent Good Poor): ")]
        grading_scale = []

        # Filling an empty list for grading
        # Will compress this algorithm later
        for grade in criteria:
            grading_scale.append(0)

        all_grades = getColumnInfo(sheet, col)
        # Use dictionary 

        for grade in all_grades:
            # pause on this...
        
            return None

    # If the number given is not an option
    else:
        print("Not an option")

def loopingAverageForColumn(sheet: openpyxl.workbook.workbook.Workbook, col_start: int, col_end: int):
    '''
    Return list of each average for the corresponding columns
    '''
    return

def getMatchedRowValues(sheet: openpyxl.workbook.workbook.Workbook, col: int, search_term):
    '''
    Grabs specified column information from each matched row
    Perfect function for getting the average of a specific class section or comments

    For future me...
    Use corresponding rows to get the list of matched rows,
    Iterate through those rows ONLY and go to specified column of those rows
    Add values of those columns to a list and return it

    :param sheet: The active sheet obj to get info from
    :param col: Column number iterate though and grab info from
    :param search: Value used when iterating through column to find the match
    '''

    rows_checking = getCorrespondingRows(sheet, col, search_term)
    
    # Dictionary holding the row and it's corresponding value
    # The key is the row

    row_values = {} 
    for row in rows_checking:
        row_values 

        return       

def splitSheet(sheet: openpyxl.workbook.workbook.Workbook):
    '''
    Splits sheet into multiple sheets based on criteria
    To be updated...
    '''
    return

def outputFile(sheet: openpyxl.workbook.workbook.Workbook, format: str):
    '''
    Creates a file of the compiled information in the format given
    Supports text and spreadsheet files

    :param sheet: The active sheet obj to get info from
    :param format: Extension of the output file, either "text" for a text file or "spread" for spreadsheet file
    '''

    match format.lower():
        case "text":
            print("Outputting to text file...")

        case "spread":
            print("Outputting to spreadsheet file...")

        case _:
            print("Incorrect format given, try again")

